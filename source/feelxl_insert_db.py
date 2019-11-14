import pymysql
import pandas as pd
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl import load_workbook

def insert_excel_to_db():
    conn = pymysql.connect(host='106.10.32.85', 
                        user='root', 
                        passwd='shsmsrpwhgdktjqjdlqslek!', 
                        db='SPST_S',
                        charset='utf8',
                        port=3306,
                        use_unicode=True)
    cursor = conn.cursor()

    df = pd.read_excel('feeling.xlsx',sheet_name='Sheet1')
    del df["일련번호"]
    del df["빈도"]
    del df["감정정도M"]
    del df["감정정도STD"]

    df.to_excel('Rfeeling.xlsx', sheet_name='Sheet1')
    print(df)

    try:
        with conn.cursor() as curs:
            sql = 'insert into FEELXL values(%s, %s)'
 
            wb = load_workbook('Rfeeling.xlsx',data_only=True)
            ws = wb['Sheet1']
 
            iter_rows = iter(ws.rows)
            next(iter_rows)
            for row in iter_rows:
                curs.execute(sql, (row[1].value, row[2].value))
            conn.commit()
    finally:
        conn.close()
        wb.close()
 
 
if __name__ == "__main__":
    #delete_all()
    insert_excel_to_db()
    #select_all()
